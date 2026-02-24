import os
import uuid
import json
import pandas as pd
import numpy as np
from flask import Flask, request, render_template, jsonify, flash, redirect, url_for, send_file
from werkzeug.utils import secure_filename
import concurrent.futures
import time
from io import BytesIO
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from datetime import datetime
from tqdm import tqdm 
from collections import Counter  # Added for analysis
import openpyxl  # For Excel export
import re
from rapidfuzz import fuzz, process  # For fuzzy name matching

app = Flask(__name__)
app.secret_key = 'threatcipher_secret_key'

EXCEL_FOLDER = 'Excel_files'
UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'csv', 'xlsx', 'xls'}
MAX_WORKERS = 4
KEYWORDS_FILE = 'ignore_keywords.json'

DEFAULT_IGNORE_KEYWORDS = [
    ' Kumar', 'mohammad', 'muhammad', 'muhammed', 'mohammed', 'ahamad', 'kumar', 'Kumar'
]
# Ignored name prefixes/common words specifically for the /check endpoint.
# --- search_by_index ported from mod/ter.py ---
def search_by_index(query, field=None):
    query = str(query).lower().strip()
    if len(query) < 5:
        return "Query must be at least 5 characters long.", pd.DataFrame(), query
    if any(keyword.lower() in query for keyword in IGNORE_KEYWORDS):
        return "Query contains ignored keyword(s).", pd.DataFrame(), query
    results = []
    seen_records = set()
    # Get all query token variations using alphanumeric tokenization
    query_tokens = tokenize_alphanumeric(query)
    for query_variant in query_tokens:
        query_hash = hash(query_variant)  # Use hash for fast lookup
        # Check if the query hash exists in the index
        if query_hash in hash_index:
            for filename, idx, col, original_value, record_id in hash_index[query_hash]:
                # Skip if we already added this record
                if record_id in seen_records:
                    continue
                if field is None or field == col:
                    # For comma-separated values, check if query matches any part exactly
                    if ',' in original_value or ' ' in original_value:
                        # Split by both commas and spaces
                        parts = []
                        # First split by commas
                        comma_parts = [part.strip().lower() for part in original_value.split(',')]
                        # Then split each comma part by spaces if needed
                        for part in comma_parts:
                            if len(part) >= 5:  # Only consider parts that are long enough
                                parts.append(part)
                            # Split by spaces
                            space_parts = [sp.strip().lower() for sp in part.split()]
                            for sp in space_parts:
                                if len(sp) >= 5:  # Only consider parts that are long enough
                                    parts.append(sp)
                        # Check all tokenized versions of parts
                        match_found = False
                        for part in parts:
                            part_tokens = tokenize_alphanumeric(part)
                            if query_variant in part_tokens:
                                match_found = True
                                break
                        if match_found:
                            row = all_excel_data[filename].iloc[idx].copy()
                            row['Source File'] = filename
                            row['Matched Term'] = query
                            row['record_id'] = record_id
                            results.append(row)
                            seen_records.add(record_id)
                    # Also check for full cell match
                    else:
                        original_tokens = tokenize_alphanumeric(original_value.lower())
                        if query_variant in original_tokens:
                            row = all_excel_data[filename].iloc[idx].copy()
                            row['Source File'] = filename
                            row['Matched Term'] = query
                            row['record_id'] = record_id
                            results.append(row)
                            seen_records.add(record_id)
    if results:
        combined_results = pd.DataFrame(results)
        return "WARNING: Match found in List!", combined_results, query
    return "No Threat, Access Granted.", pd.DataFrame(), query
CHECK_IGNORE_WORDS = {
    # Mohammed variants
    'mohammad', 'mohammed', 'mohamed', 'muhammad', 'muhammed', 'muhamad',
    'mohamad', 'mohammod', 'mahammad', 'mehmed', 'mehmet',
    # Common filler names — NOT 'ali' (it appears in real names like wali, somali)
    'kumar', 'kumari', 'ahamad', 'ahmad', 'ahmed',
    'bin', 'binti', 'binte', 'abd', 'abdul', 'abdur', 'abdel',
    # Titles/prefixes
    'md', 'md.', 'mr', 'mr.', 'mrs', 'mrs.', 'ms', 'ms.', 'dr', 'dr.',
    'haji', 'hajj', 'haj', 'hj', 'sheikh', 'shaikh', 'shaykh',
}

def _strip_check_ignored(query):
    """
    For /check endpoint: split query into tokens, drop any token that is
    in CHECK_IGNORE_WORDS or shorter than 3 characters, return remaining
    tokens joined. If nothing remains, return the original query unchanged
    so we don't search with an empty string.
    Example: 'MOHAMMED RIFAN' -> 'rifan' (only meaningful token kept)
             'MOHAMMED ALI'   -> 'ali'   ('ali' is NOT ignored, it is a real name)
             'KUMAR RAJESH'   -> 'rajesh'
    """
    tokens = query.lower().strip().split()
    kept = [t for t in tokens if len(t) >= 3 and t not in CHECK_IGNORE_WORDS]
    if not kept:
        # All tokens were ignored — fall back to longest non-ignored token
        non_ignored = [t for t in tokens if t not in CHECK_IGNORE_WORDS and len(t) >= 3]
        return ' '.join(non_ignored) if non_ignored else query.lower().strip()
    return ' '.join(kept)

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

NAME_VARIANTS = ['name', 'names', 'first name', 'second name', 'third name']
REQUIRED_COLUMNS = ['NIC', 'Passport', 'Address1', 'Phone11', 'Birth Day', 'First Name', 'Last Name']

# Columns to index for the /check endpoint only - keeps it fast and precise
CHECK_INDEX_COLUMNS = {
    'name', 'original script name', 'aliases',
    'reference number', 'document number', 'dl/ passport no.', 'nic no.', 'id',
    'date of birth', 'place of birth', 'nationality', 'citizenship',
    'address', 'address (sri lanka)', 'address (foreign)',
}

CHECK_NAME_PARTS = {'first name', 'second name', 'third name'}

def load_ignore_keywords():
    if os.path.exists(KEYWORDS_FILE):
        with open(KEYWORDS_FILE, 'r') as f:
            try:
                keywords = json.load(f)
                return keywords if keywords else DEFAULT_IGNORE_KEYWORDS
            except json.JSONDecodeError:
                return DEFAULT_IGNORE_KEYWORDS
    else:
        with open(KEYWORDS_FILE, 'w') as f:
            json.dump(DEFAULT_IGNORE_KEYWORDS, f)
        return DEFAULT_IGNORE_KEYWORDS

def save_ignore_keywords(keywords):
    with open(KEYWORDS_FILE, 'w') as f:
        json.dump(keywords, f)

IGNORE_KEYWORDS = load_ignore_keywords()

def tokenize_alphanumeric(value):
    """    
    Returns a set of normalized variations for indexing/searching
    """
    if not isinstance(value, str) or len(value) < 5:
        return {value}
    
    tokens = {value}  
    
    # This will match transitions like: a1, 1a, ab12, 12ab
    spaced_version = re.sub(r'([a-zA-Z])(\d)', r'\1 \2', value)
    spaced_version = re.sub(r'(\d)([a-zA-Z])', r'\1 \2', spaced_version)  
    
    if spaced_version != value:
        tokens.add(spaced_version)
    
    return tokens

def load_all_excel_files(folder_path):
    print("Loading excel files and building hash-based search index...")
    start_time = time.time()
    excel_data = {}
    hash_index = {}
    check_hash_index = {}   # Focused index for /check endpoint (specific columns only)
    record_id_map = {}

    files = [f for f in os.listdir(folder_path) if f.endswith(('.xlsx', '.xls'))]
    for filename in tqdm(files, desc="Loading Excel Files"):
        file_path = os.path.join(folder_path, filename)
        try:
            df = pd.read_excel(file_path)
            df = df.astype(str).apply(lambda x: x.str.strip().str.lower())
            df.fillna('', inplace=True)
            name_cols = [col for col in df.columns if col.lower() in NAME_VARIANTS]
            if name_cols:
                df['Name'] = df[name_cols].agg(' '.join, axis=1).str.strip()

            df['record_id'] = [str(uuid.uuid4()) for _ in range(len(df))]
            excel_data[filename] = df

            for idx in tqdm(range(len(df)), desc=f"Indexing {filename}", leave=False):
                row = df.iloc[idx]
                record_id = row['record_id']
                record_id_map[record_id] = (filename, idx)

                # --- Build focused check_hash_index for /check endpoint ---
                # Combine first/second/third name into a single full name and index it
                name_parts = []
                for name_col in ['first name', 'second name', 'third name']:
                    matched_col = next((c for c in df.columns if c.lower() == name_col), None)
                    if matched_col:
                        part = str(row[matched_col]).strip()
                        if part and part.lower() not in ('nan', 'none', ''):
                            name_parts.append(part)
                if name_parts:
                    full_name = ' '.join(name_parts).strip().lower()
                    if len(full_name) >= 3:
                        # Index full name
                        for ft in tokenize_alphanumeric(full_name):
                            fh = hash(ft)
                            if fh not in check_hash_index:
                                check_hash_index[fh] = []
                            check_hash_index[fh].append((filename, idx, 'full name', full_name, record_id))
                        # Index each individual name token too
                        for part in full_name.split():
                            if len(part) >= 3:
                                for pt in tokenize_alphanumeric(part):
                                    ph = hash(pt)
                                    if ph not in check_hash_index:
                                        check_hash_index[ph] = []
                                    check_hash_index[ph].append((filename, idx, 'full name', full_name, record_id))

                for col in df.columns:
                    value = row[col]
                    if isinstance(value, str) and len(value) >= 5:
                        # --- Index other CHECK_INDEX_COLUMNS (excluding name parts, handled above) ---
                        if col.lower() in CHECK_INDEX_COLUMNS and col.lower() not in CHECK_NAME_PARTS:
                            chk_tokens = tokenize_alphanumeric(value)
                            for chk_token in chk_tokens:
                                chk_hash = hash(chk_token)
                                if chk_hash not in check_hash_index:
                                    check_hash_index[chk_hash] = []
                                check_hash_index[chk_hash].append((filename, idx, col, value, record_id))
                                # Also index individual space-separated parts
                                if ' ' in chk_token:
                                    for part in chk_token.split():
                                        if len(part) >= 5:
                                            for pt in tokenize_alphanumeric(part):
                                                ph = hash(pt)
                                                if ph not in check_hash_index:
                                                    check_hash_index[ph] = []
                                                check_hash_index[ph].append((filename, idx, col, value, record_id))

                        # --- Build full hash_index for all other endpoints ---
                        tokens = tokenize_alphanumeric(value)
                        for token in tokens:
                            token_hash = hash(token)
                            if token_hash not in hash_index:
                                hash_index[token_hash] = []
                            hash_index[token_hash].append((filename, idx, col, value, record_id))
                            # Index comma-separated parts
                            if ',' in token:
                                parts = [part.strip() for part in token.split(',')]
                                for part in parts:
                                    if len(part) >= 5:
                                        part_tokens = tokenize_alphanumeric(part)
                                        for part_token in part_tokens:
                                            part_hash = hash(part_token)
                                            if part_hash not in hash_index:
                                                hash_index[part_hash] = []
                                            hash_index[part_hash].append((filename, idx, col, value, record_id))
                                            # Also index space-separated subparts within comma-separated parts
                                            if ' ' in part_token:
                                                subparts = [subpart.strip() for subpart in part_token.split()]
                                                for subpart in subparts:
                                                    if len(subpart) >= 5:
                                                        subpart_tokens = tokenize_alphanumeric(subpart)
                                                        for sp_token in subpart_tokens:
                                                            subpart_hash = hash(sp_token)
                                                            if subpart_hash not in hash_index:
                                                                hash_index[subpart_hash] = []
                                                            hash_index[subpart_hash].append((filename, idx, col, value, record_id))
                            # Directly index space-separated parts
                            elif ' ' in token:
                                parts = [part.strip() for part in token.split()]
                                for part in parts:
                                    if len(part) >= 5:
                                        part_tokens = tokenize_alphanumeric(part)
                                        for part_token in part_tokens:
                                            part_hash = hash(part_token)
                                            if part_hash not in hash_index:
                                                hash_index[part_hash] = []
                                            hash_index[part_hash].append((filename, idx, col, value, record_id))
        except Exception as e:
            print(f"Error loading {filename}: {e}")

    print(f"Finished loading {len(excel_data)} files and building index with {len(hash_index)} keys ({len(check_hash_index)} check-specific keys) in {time.time() - start_time:.2f} seconds")
    return excel_data, hash_index, check_hash_index, record_id_map

all_excel_data, hash_index, check_hash_index, record_id_map = load_all_excel_files(EXCEL_FOLDER)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def fuzzy_search_check(query, threshold=80):
    """
    Fuzzy search for /check endpoint — scans only CHECK_INDEX_COLUMNS.
    Strips CHECK_IGNORE_WORDS from query before scoring.
    Scores against full cell value AND each individual token (min 3 chars).
    Cell tokens that are in CHECK_IGNORE_WORDS are also skipped.
    """
    from difflib import SequenceMatcher

    # Strip ignored words from query (mohammed ali -> ali, or just meaningful part)
    cleaned_query = _strip_check_ignored(query)
    query_lower = cleaned_query.lower().strip()

    if len(query_lower) < 3:
        return []

    results = []
    seen_records = set()

    for filename, df in all_excel_data.items():
        for idx in range(len(df)):
            row = df.iloc[idx]
            record_id = row.get('record_id', '')
            if record_id in seen_records:
                continue

            best_score = 0
            best_matched_col = ''
            best_matched_val = ''

            # Check combined full name (first + second + third name)
            name_parts = []
            for name_col in ['first name', 'second name', 'third name']:
                matched_col = next((c for c in df.columns if c.lower() == name_col), None)
                if matched_col:
                    part = str(row[matched_col]).strip().lower()
                    if part and part not in ('nan', 'none', ''):
                        name_parts.append(part)
            if name_parts:
                full_name = ' '.join(name_parts).strip()
                cleaned_cell_name = _strip_check_ignored(full_name)
                if len(cleaned_cell_name) >= 3:
                    score_full = SequenceMatcher(None, query_lower, cleaned_cell_name).ratio() * 100
                    if score_full > best_score:
                        best_score = score_full
                        best_matched_col = 'full name'
                        best_matched_val = full_name
                # Token-by-token (min 3 chars, skip ignored, short-token guard)
                for token in full_name.split():
                    if len(token) < 3 or token in CHECK_IGNORE_WORDS:
                        continue
                        # Short query guard: short queries only match short full names (<=2 meaningful words)
                    if len(query_lower) <= 4:
                        cleaned_fn = _strip_check_ignored(full_name)
                        if len(cleaned_fn.split()) > 2:
                            continue
                    s = SequenceMatcher(None, query_lower, token).ratio() * 100
                    if s > best_score:
                        best_score = s
                        best_matched_col = 'full name'
                        best_matched_val = full_name

            # Check other focused columns (skip name parts, already handled above)
            for col in df.columns:
                if col.lower() not in CHECK_INDEX_COLUMNS or col.lower() in CHECK_NAME_PARTS:
                    continue
                cell_value = str(row[col]).lower().strip()
                if len(cell_value) < 3:
                    continue

                # Score against full cell value (after stripping ignored words)
                cleaned_cell = _strip_check_ignored(cell_value)
                if len(cleaned_cell) >= 3:
                    score_full = SequenceMatcher(None, query_lower, cleaned_cell).ratio() * 100
                    if score_full > best_score:
                        best_score = score_full
                        best_matched_col = col
                        best_matched_val = cell_value

                # Token-by-token (min 3 chars, skip ignored, short-query guard)
                for token in cell_value.split():
                    if len(token) < 3 or token in CHECK_IGNORE_WORDS:
                        continue
                    # Short query guard: short queries only accepted if the cleaned
                    # cell itself is short (<=2 meaningful words)
                    if len(query_lower) <= 4:
                        cleaned_cv = _strip_check_ignored(cell_value)
                        if len(cleaned_cv.split()) > 2:
                            continue
                    score_token = SequenceMatcher(None, query_lower, token).ratio() * 100
                    if score_token > best_score:
                        best_score = score_token
                        best_matched_col = col
                        best_matched_val = cell_value

                if best_score >= threshold:
                    break

            if best_score >= threshold:
                seen_records.add(record_id)
                record_dict = row.to_dict()
                record_dict['Source File'] = filename
                record_dict['Matched Term'] = query
                record_dict['Matched Column'] = best_matched_col
                record_dict['Matched Value'] = best_matched_val
                results.append((record_dict, round(best_score, 2)))

    return results

def search_by_check_index(query):
    """
    Direct hash lookup using check_hash_index (specific columns only).
    Used exclusively by the /check endpoint.
    Strips CHECK_IGNORE_WORDS tokens from query before searching.
    Minimum token length: 3 characters.
    """
    # Strip ignored words (mohammed, kumar, etc.) from query first
    cleaned_query = _strip_check_ignored(query)
    query_lower = cleaned_query.lower().strip()

    if len(query_lower) < 3:
        return pd.DataFrame()

    results = []
    seen_records = set()
    # Tokenize and drop any remaining short or ignored tokens
    query_tokens = {t for t in tokenize_alphanumeric(query_lower)
                    if len(t) >= 3 and t not in CHECK_IGNORE_WORDS}

    for query_variant in query_tokens:
        query_hash = hash(query_variant)
        if query_hash in check_hash_index:
            for filename, idx, col, original_value, record_id in check_hash_index[query_hash]:
                if record_id in seen_records:
                    continue

                orig_lower = original_value.lower()

                # Verify it's a real match not just a hash collision
                orig_tokens = tokenize_alphanumeric(orig_lower)
                cell_parts = set()
                cell_parts.update(orig_tokens)
                for t in orig_lower.split():
                    cell_parts.update(tokenize_alphanumeric(t))

                if query_variant not in cell_parts:
                    continue

                # Short token guard: if cleaned query is short (<=4 chars),
                # only accept if the entire cell value (after stripping ignored words)
                # is also short (<=2 words). Prevents 'ali' matching
                # 'abdul salam hanafi ali mardan' but allows it to match 'ali' or 'wali ali'.
                if len(query_variant) <= 4:
                    cleaned_cell = _strip_check_ignored(orig_lower)
                    if len(cleaned_cell.split()) > 2:
                        continue

                row = all_excel_data[filename].iloc[idx].copy()
                row['Source File'] = filename
                # Log which column and value triggered the match
                row['Matched Term'] = query
                row['Matched Column'] = col
                row['Matched Value'] = original_value
                results.append(row)
                seen_records.add(record_id)

    return pd.DataFrame(results) if results else pd.DataFrame() 

def process_upload_file(file_path):
    file_ext = file_path.rsplit('.', 1)[1].lower()
    try:
        if file_ext == 'csv':
            return process_csv_file(file_path)
        elif file_ext in ['xlsx', 'xls']:
            return process_excel_file(file_path)
        else:
            return None, f"Unsupported file format: {file_ext}"
    except Exception as e:
        return None, f"Error processing file: {str(e)}"

def process_csv_file(file_path):
    try:
        csv_chunks = pd.read_csv(file_path, chunksize=10000)
        processed_chunks = []
        
        for chunk in tqdm(csv_chunks, desc="Processing CSV Chunks"):
            chunk.columns = chunk.columns.str.strip()
            for col in REQUIRED_COLUMNS:
                if col not in chunk.columns:
                    chunk[col] = None
            for col in REQUIRED_COLUMNS:
                if col in chunk.columns:
                    chunk[col] = chunk[col].fillna("None").astype(str).str.strip()
            chunk['Full Name'] = chunk['First Name'] + ' ' + chunk['Last Name']
            processed_chunks.append(chunk)
            
        df = pd.concat(processed_chunks)
        return df, None
    except Exception as e:
        return None, f"Error processing CSV file: {str(e)}"

def process_excel_file(file_path):
    try:
        df = pd.read_excel(file_path)
        df.columns = df.columns.str.strip()
        
        for col in REQUIRED_COLUMNS:
            if col not in df.columns:
                df[col] = None
                
        for col in REQUIRED_COLUMNS:
            if col in df.columns:
                df[col] = df[col].fillna("None").astype(str).str.strip()
                
        df['Full Name'] = df['First Name'] + ' ' + df['Last Name']
        
        return df, None
    except Exception as e:
        return None, f"Error processing Excel file: {str(e)}"

# Add a new function for the multi-stage search
def search_from_uploaded_data_staged(df, stage=1):
    start_time = time.time()
    print(f"Processing {len(df)} records on GPU with stage {stage}...")
    
    # Create batches
    num_batches = min(MAX_WORKERS, max(1, len(df) // 5000))
    batch_size = max(1, len(df) // num_batches)
    batches = [(i, df.iloc[i:i+batch_size], stage) for i in range(0, len(df), batch_size)]
    
    grouped_results = {}
    csv_entries = {}
    matched_terms = []
    
    with concurrent.futures.ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        future_to_batch = {executor.submit(search_batch_gpu_staged, batch): batch for batch in batches}
        for future in tqdm(concurrent.futures.as_completed(future_to_batch), total=len(batches), desc=f"Processing Batches (Stage {stage})"):
            batch_results, batch_entries, batch_terms = future.result()
            
            for idx, df in batch_results.items():
                grouped_results[idx] = df
            for idx, entry in batch_entries.items():
                csv_entries[idx] = entry
            matched_terms.extend(batch_terms)
    
    print(f"Stage {stage} data processing completed in {time.time() - start_time:.2f} seconds. Found {len(grouped_results)} matches.")
    
    if grouped_results:
        return f"WARNING: Matches found for uploaded entries (Stage {stage})!", grouped_results, csv_entries, matched_terms
    return f"No matches found for any uploaded entries (Stage {stage}).", {}, {}, []

def search_batch_gpu_staged(batch_data):
    batch_idx, batch_df, stage = batch_data
    results = {}
    csv_entries = {}
    matched_terms = []
    
    # Process each row in the batch
    for idx, row in batch_df.iterrows():
        # Store CSV entry data
        csv_entry = {
            'MID': str(row.get('MID', 'None')),
            'NIC': str(row['NIC']),
            'Passport': str(row['Passport']),
            'Phone11': str(row['Phone11']),
            'Birth Day': str(row['Birth Day']),
            'First Name': str(row['First Name']),
            'Last Name': str(row['Last Name']),
            'Index': idx
        }
        csv_entries[idx] = csv_entry
        
        # Define search terms based on stage
        search_terms = []
        fields = []
        
        # Stage 1: ID and contact info search
        if stage == 1:
            for field in ['NIC', 'Passport', 'Phone11']:
                if field in row and row[field] and len(str(row[field])) >= 5:
                    term = str(row[field]).lower()
                    if not any(keyword.lower() in term for keyword in IGNORE_KEYWORDS):
                        search_terms.append(term)
                        fields.append(field)
        
        # Stage 2: First Name OR Last Name
        elif stage == 2:
            for field in ['First Name', 'Last Name']:
                if field in row and row[field] and len(str(row[field])) >= 5:
                    term = str(row[field]).lower()
                    if not any(keyword.lower() in term for keyword in IGNORE_KEYWORDS):
                        search_terms.append(term)
                        fields.append(field)
        
        # Stage 3: First Name AND Last Name combined
        elif stage == 3:
            first_name = str(row.get('First Name', '')).lower()
            last_name = str(row.get('Last Name', '')).lower()
            
            if first_name and last_name and len(first_name) >= 2 and len(last_name) >= 2:
                if not any(keyword.lower() in first_name for keyword in IGNORE_KEYWORDS) and \
                   not any(keyword.lower() in last_name for keyword in IGNORE_KEYWORDS):
                    # Combine first and last name
                    term = f"{first_name} {last_name}".strip()
                    if len(term) >= 5:
                        search_terms.append(term)
                        fields.append('Full Name')
        
        # Avoid duplicate matches by tracking matched record IDs
        matched_record_ids = set()
        
        # Perform search for the terms
        for i, term in enumerate(search_terms):
            field = fields[i]
            message, matches_df, matched_term = search_by_index(term)
            
            if not matches_df.empty:
                # For Stage 3, verify both first and last name match
                if stage == 3:
                    # Additional verification for stage 3
                    first_name_lower = row.get('First Name', '').lower()
                    last_name_lower = row.get('Last Name', '').lower()
                    
                    # Filter matches to ensure both first and last name appear
                    valid_matches = []
                    for _, match_row in matches_df.iterrows():
                        match_name = match_row.get('Name', '').lower()
                        if first_name_lower in match_name and last_name_lower in match_name:
                            # Check if we haven't already matched this record
                            record_id = match_row.get('record_id')
                            if record_id not in matched_record_ids:
                                matched_record_ids.add(record_id)
                                valid_matches.append(match_row)
                    
                    if valid_matches:
                        matches_df = pd.DataFrame(valid_matches)
                    else:
                        continue  # Skip if no valid matches
                else:
                    # Filter out already matched records
                    new_matches = []
                    for _, match_row in matches_df.iterrows():
                        record_id = match_row.get('record_id')
                        if record_id not in matched_record_ids:
                            matched_record_ids.add(record_id)
                            new_matches.append(match_row)
                    
                    if new_matches:
                        matches_df = pd.DataFrame(new_matches)
                    else:
                        continue  # Skip if all matches are duplicates
                
                matches_df['CSV Source'] = True
                matches_df['Search Field'] = field
                
                # Initialize results for this row if needed
                if idx not in results:
                    results[idx] = matches_df
                else:
                    # Append new matches to existing ones
                    results[idx] = pd.concat([results[idx], matches_df], ignore_index=True)
                
                matched_terms.append(matched_term)
    
    return results, csv_entries, matched_terms

def search_from_uploaded_data(df):
    start_time = time.time()
    print(f"Processing {len(df)} records on GPU...")
    
    num_batches = min(MAX_WORKERS, max(1, len(df) // 5000))
    batch_size = max(1, len(df) // num_batches)
    batches = [(i, df.iloc[i:i+batch_size]) for i in range(0, len(df), batch_size)]
    
    grouped_results = {}
    csv_entries = {}
    matched_terms = []
    
    with concurrent.futures.ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        future_to_batch = {executor.submit(search_batch_gpu_staged, batch): batch for batch in batches}
        for future in tqdm(concurrent.futures.as_completed(future_to_batch), total=len(batches), desc="Processing Batches"):
            batch_results, batch_entries, batch_terms = future.result()
            
            for idx, df in batch_results.items():
                grouped_results[idx] = df
            for idx, entry in batch_entries.items():
                csv_entries[idx] = entry
            matched_terms.extend(batch_terms)
    
    print(f"Data processing completed in {time.time() - start_time:.2f} seconds. Found {len(grouped_results)} matches.")
    
    if grouped_results:
        return "WARNING: Matches found for uploaded entries!", grouped_results, csv_entries, matched_terms
    return "No matches found for any uploaded entries.", {}, {}, []

import csv

def log_to_csv(input_data, output_data, folder_name="Auto_Check"):
    """
    Log /check endpoint requests and responses to a daily CSV file.
    Includes match_type (direct/fuzzy) and fuzzy_confidence per entry.
    """
    try:
        # Always resolve folder relative to this script, not Flask's cwd
        base_dir = os.path.dirname(os.path.abspath(__file__))
        folder_path = os.path.join(base_dir, folder_name)
        os.makedirs(folder_path, exist_ok=True)

        # Generate filename with current date
        current_date = datetime.now().strftime("%Y-%m-%d")
        csv_filename = os.path.join(folder_path, f"check_log_{current_date}.csv")
        print(f"[log_to_csv] Writing to: {csv_filename}")

        # Check if file exists to determine if we need headers
        file_exists = os.path.isfile(csv_filename)

        # Pull matches list
        matches = output_data.get('matches', [])

        # Summarise match types and fuzzy confidences
        match_types = sorted({m.get('match_type', 'direct') for m in matches if isinstance(m, dict)})
        fuzzy_confidences = [
            str(m.get('fuzzy_confidence', ''))
            for m in matches
            if isinstance(m, dict) and m.get('match_type') == 'fuzzy'
        ]

        # Build a compact per-match summary: name|source|match_type|confidence
        match_summary_rows = []
        for m in matches:
            if isinstance(m, dict):
                rid = m.get('record_id', 'N/A')
                mtype = m.get('match_type', 'direct')
                conf = m.get('fuzzy_confidence', 100.0)
                name = m.get('Name', m.get('name', ''))
                src = m.get('Source File', '')
                match_summary_rows.append(f"{name}|{src}|{mtype}|{conf}")

        # These fieldnames MUST exactly match the log_entry keys below
        fieldnames = [
            'timestamp',
            'input_keyword',
            'query_used',
            'output_status',
            'result_message',
            'matched_term',
            'matches_count',
            'match_types',
            'fuzzy_confidences',
            'match_summary',
            'matches_details',
            'error_message',
        ]

        log_entry = {
            'timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            'input_keyword': input_data.get('keyword', ''),
            'query_used': output_data.get('query_used', input_data.get('keyword', '')),
            'output_status': output_data.get('status', ''),
            'result_message': output_data.get('result_message', ''),
            'matched_term': output_data.get('matched_term', ''),
            'matches_count': len(matches),
            'match_types': ', '.join(match_types) if match_types else '',
            'fuzzy_confidences': ', '.join(fuzzy_confidences) if fuzzy_confidences else '',
            'match_summary': ' ;; '.join(match_summary_rows) if match_summary_rows else '',
            'matches_details': json.dumps(matches, ensure_ascii=False),
            'error_message': output_data.get('error', ''),
        }

        with open(csv_filename, 'a', newline='', encoding='utf-8') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            if not file_exists:
                writer.writeheader()
            writer.writerow(log_entry)

        print(f"[log_to_csv] Logged to {csv_filename} — status={log_entry['output_status']}, matches={log_entry['matches_count']}, types={log_entry['match_types']}")

    except Exception as e:
        print(f"[log_to_csv] ERROR: {type(e).__name__}: {str(e)}")

# Modify the route handler to support stages
@app.route('/', methods=['GET', 'POST'])
def index():
    global IGNORE_KEYWORDS
    result_message = ""
    result_data = {}
    csv_entries = {}
    matched_terms = []
    
    if request.method == 'POST':
        if 'query' in request.form:
            query = request.form.get('query', '').strip()
            field = request.form.get('field', '')
            if field == "All Fields" or not field:
                field = None

            result_message, matches, matched_term = search_by_index(query, field)
            if not matches.empty:
                result_data = {'manual': matches.to_dict(orient='records')}
                matched_terms = [matched_term]
        
        elif 'uploadFile' in request.files:
            file = request.files['uploadFile']
            if file and allowed_file(file.filename):
                # Get search stage
                search_stage = int(request.form.get('searchStage', '1'))
                
                # Get ignore keywords
                ignore_keywords_input = request.form.get('ignoreKeywords', '')
                if ignore_keywords_input:
                    new_keywords = [kw.strip().lower() for kw in ignore_keywords_input.split(',') if kw.strip()]
                    IGNORE_KEYWORDS = list(set(DEFAULT_IGNORE_KEYWORDS + new_keywords))
                    save_ignore_keywords(IGNORE_KEYWORDS)
                else:
                    IGNORE_KEYWORDS = load_ignore_keywords()
                
                filename = secure_filename(file.filename)
                file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                file.save(file_path)
                
                start_time = time.time()
                df, error = process_upload_file(file_path)
                if error:
                    result_message = error
                else:
                    # Call the staged search function with the selected stage
                    result_message, grouped_results, csv_entries, matched_terms = search_from_uploaded_data_staged(df, search_stage)
                    result_data = grouped_results
                    print(f"Total processing time: {time.time() - start_time:.2f} seconds")
    
    common_fields = set()
    for df in all_excel_data.values():
        common_fields.update(df.columns)
    fields = ["All Fields"] + sorted(list(common_fields))
    
    return render_template('index.html', 
                         result_message=result_message, 
                         result_data=result_data, 
                         csv_entries=csv_entries,
                         fields=fields,
                         matched_terms=matched_terms,
                         ignore_keywords=','.join(IGNORE_KEYWORDS))

@app.route('/check', methods=['POST'])
def check():
    # Capture input data for logging
    input_data = {
        'method': request.method,
        'is_json': request.is_json
    }

    if request.is_json:
        data = request.get_json()
        query = data.get('keyword', '').strip()
    else:
        query = request.form.get('keyword', '').strip()

    input_data['keyword'] = query

    if not query:
        error_response = {'error': 'No keyword provided', 'status': 'failure'}
        log_to_csv(input_data, error_response)
        return jsonify(error_response), 400

    all_matches = []

    # Strip ignored words from query for /check (mohammed ali -> ali)
    cleaned_query = _strip_check_ignored(query)

    # --- Step 1: Direct hash lookup using check_hash_index (focused columns only) ---
    matches = search_by_check_index(query)  # internally strips ignored words
    matched_term = cleaned_query

    if not matches.empty:
        for record in matches.to_dict(orient='records'):
            record['match_type'] = 'direct'
            record['fuzzy_confidence'] = 100.0
            all_matches.append(record)

    # --- Step 2: Fuzzy match (focused columns only, >=80% threshold) ---
    # Only run if no direct matches found, to avoid duplicates
    if not all_matches:
        fuzzy_results = fuzzy_search_check(query, threshold=80)
        for record_dict, score in fuzzy_results:
            record_dict['match_type'] = 'fuzzy'
            record_dict['fuzzy_confidence'] = score
            all_matches.append(record_dict)

    # Determine final result
    if all_matches:
        final_message = "WARNING: Match found in List!"
        status = 'success'
    else:
        final_message = "No Threat, Access Granted."
        status = 'no_match'
        matched_term = query

    response = {
        'result_message': final_message,
        'matched_term': matched_term,
        'matches': all_matches,
        'status': status,
        'query_used': query,           # Original input as received
        'query_searched': cleaned_query  # After stripping ignored words
    }

    # Log the request and response
    log_to_csv(input_data, response)

    return jsonify(response)

@app.route('/get_record_details', methods=['POST'])
def get_record_details():
    data = request.get_json()
    record_id = data.get('record_id')
    
    if not record_id:
        return jsonify({'error': 'Missing record_id'}), 400
    
    if record_id not in record_id_map:
        return jsonify({'error': 'Record not found'}), 404
    
    filename, idx = record_id_map[record_id]
    record = all_excel_data[filename].iloc[idx].to_dict()
    record['Source File'] = filename
    record['Matched Term'] = data.get('matched_term', '')
    
    return jsonify(record)

@app.route('/generate_pdf', methods=['POST'])
def generate_pdf():
    data = request.get_json()
    result_message = data.get('result_message', '')
    result_data = data.get('result_data', {})
    csv_entries = data.get('csv_entries', {})
    matched_terms = data.get('matched_terms', [])

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name='CustomHeading', fontSize=16, textColor='#00c3ff', fontName='Helvetica-Bold'))
    styles.add(ParagraphStyle(name='CustomNormal', fontSize=10, textColor='#00c3ff', fontName='Helvetica'))
    story = []

    story.append(Paragraph(f"ThreatCipher Report - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['CustomHeading']))
    story.append(Spacer(1, 12))

    alert_color = '#ff4d4d' if 'WARNING' in result_message else '#4dff4d'
    story.append(Paragraph(f"<font color='{alert_color}'>{result_message}</font>", styles['CustomNormal']))
    story.append(Spacer(1, 12))

    if result_data.get('manual'):
        story.append(Paragraph("Manual Search Results", styles['CustomHeading']))
        for record in result_data['manual']:
            story.append(Paragraph(f"Name: {record.get('Name', 'Unknown').capitalize()}", styles['CustomNormal']))
            story.append(Paragraph(f"Source: {record.get('Source File')}", styles['CustomNormal']))
            story.append(Paragraph(f"Match: {record.get('Matched Term')}", styles['CustomNormal']))
            story.append(Paragraph(f"Record ID: {record.get('record_id', 'N/A')}", styles['CustomNormal']))
            story.append(Spacer(1, 6))
    elif result_data:
        story.append(Paragraph("Upload Results", styles['CustomHeading']))
        for idx, matches in result_data.items():
            entry = csv_entries.get(str(idx), {})
            story.append(Paragraph(f"Entry #{int(idx) + 1}", styles['CustomNormal']))
            story.append(Paragraph(f"MID: {entry.get('MID', 'N/A') if entry.get('MID')  not in ['None', 'nan'] else 'N/A'}", styles['CustomNormal']))
            story.append(Paragraph(f"Name: {entry.get('First Name', '')} {entry.get('Last Name', '')}", styles['CustomNormal']))
            story.append(Paragraph(f"Phone: {entry.get('Phone11', 'N/A') if entry.get('Phone11') not in ['None', 'nan'] else 'N/A'}", styles['CustomNormal']))
            story.append(Paragraph(f"DOB: {entry.get('Birth Day', 'N/A') if entry.get('Birth Day') not in ['None', 'nan'] else 'N/A'}", styles['CustomNormal']))
            story.append(Spacer(1, 6))
            for record in matches:
                story.append(Paragraph(f"Name: {record.get('Name', 'Unknown').capitalize()}", styles['CustomNormal']))
                story.append(Paragraph(f"Source: {record.get('Source File')}", styles['CustomNormal']))
                story.append(Paragraph(f"Match: {record.get('Matched Term')}", styles['CustomNormal']))
                story.append(Paragraph(f"Record ID: {record.get('record_id', 'N/A')}", styles['CustomNormal']))
                story.append(Spacer(1, 6))
            if idx != list(result_data.keys())[-1]:
                story.append(Spacer(1, 12))
    else:
        # No matches found
        story.append(Paragraph("No threats found in List.", styles['CustomNormal']))

    doc.build(story)
    buffer.seek(0)

    return send_file(
        buffer,
        download_name=f"ThreatCipher_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf",
        as_attachment=True,
        mimetype='application/pdf'
    )

@app.route('/generate_excel', methods=['POST'])
def generate_excel():
    data = request.get_json()
    result_message = data.get('result_message', '')
    result_data = data.get('result_data', {})
    csv_entries = data.get('csv_entries', {})
    matched_terms = data.get('matched_terms', [])

    # Create a new workbook
    wb = openpyxl.Workbook()
    
    # Create summary sheet
    summary_sheet = wb.active
    summary_sheet.title = "Summary"
    
    # Add summary information
    summary_sheet['A1'] = "ThreatCipher Security Report"
    summary_sheet['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    summary_sheet['A4'] = f"Result: {result_message}"
    
    # Style formatting
    summary_sheet['A1'].font = openpyxl.styles.Font(size=16, bold=True)
    summary_sheet['A4'].font = openpyxl.styles.Font(bold=True)
    if 'WARNING' in result_message:
        summary_sheet['A4'].font = openpyxl.styles.Font(bold=True, color="FF0000")
    elif 'No Threat' in result_message or 'No matches' in result_message:
        summary_sheet['A4'].font = openpyxl.styles.Font(bold=True, color="00FF00")

    # Add analytics section in the summary sheet
    summary_sheet['A6'] = "ANALYTICS SUMMARY"
    summary_sheet['A6'].font = openpyxl.styles.Font(size=14, bold=True)
    
    # Calculate total entries and matches
    total_entries = 0
    total_matches = 0
    
    if 'manual' in result_data:
        total_entries = 1  # Manual search is a single entry
        total_matches = len(result_data['manual'])
    else:
        total_entries = len(csv_entries)
        total_matches = sum(len(matches) for matches in result_data.values())
    
    # Add analytics data
    summary_sheet['A8'] = "Total Entries:"
    summary_sheet['B8'] = total_entries
    summary_sheet['A8'].font = openpyxl.styles.Font(bold=True)
    
    summary_sheet['A9'] = "Total Matches:"
    summary_sheet['B9'] = total_matches
    summary_sheet['A9'].font = openpyxl.styles.Font(bold=True)
    
    # Calculate match counts per term
    term_counts = Counter(matched_terms)
    
    # Add term frequency table
    summary_sheet['A11'] = "MATCHED TERMS FREQUENCY"
    summary_sheet['A11'].font = openpyxl.styles.Font(size=12, bold=True)
    
    summary_sheet['A12'] = "Term"
    summary_sheet['B12'] = "Count"
    summary_sheet['A12'].font = openpyxl.styles.Font(bold=True)
    summary_sheet['B12'].font = openpyxl.styles.Font(bold=True)
    
    row = 13
    for term, count in term_counts.most_common():
        summary_sheet[f'A{row}'] = term
        summary_sheet[f'B{row}'] = count
        row += 1

    # Create Analytics sheet
    analytics_sheet = wb.create_sheet(title="Analytics")
    
    # Headers for analytics
    analytics_sheet['A1'] = "ThreatCipher Analytics"
    analytics_sheet['A1'].font = openpyxl.styles.Font(size=16, bold=True)
    
    # Entry-wise match counts
    analytics_sheet['A3'] = "ENTRY-WISE MATCH ANALYSIS"
    analytics_sheet['A3'].font = openpyxl.styles.Font(size=14, bold=True)
    
    analytics_sheet['A4'] = "Entry"
    analytics_sheet['B4'] = "Match Count"
    analytics_sheet['C4'] = "Matched Terms"
    analytics_sheet['A4'].font = openpyxl.styles.Font(bold=True)
    analytics_sheet['B4'].font = openpyxl.styles.Font(bold=True)
    analytics_sheet['C4'].font = openpyxl.styles.Font(bold=True)
    
    # Fill entry-wise data
    row = 5
    if 'manual' in result_data:
        # For manual search
        analytics_sheet[f'A{row}'] = "Manual Search"
        analytics_sheet[f'B{row}'] = len(result_data['manual'])
        analytics_sheet[f'C{row}'] = ", ".join(set(matched_terms))
    else:
        # For file uploads
        for idx, matches in result_data.items():
            analytics_sheet[f'A{row}'] = f"Entry #{int(idx) + 1}"
            analytics_sheet[f'B{row}'] = len(matches)
            
            # Get matched terms for this entry
            entry_terms = []
            for match in matches:
                term = match.get('Matched Term', '')
                if term:
                    entry_terms.append(term)
            
            analytics_sheet[f'C{row}'] = ", ".join(set(entry_terms))
            row += 1

    # Add data based on search type
    if result_data.get('manual'):
        # Create sheet for manual search results
        manual_sheet = wb.create_sheet(title="Manual Search Results")
        
        # Add headers
        headers = ["Name", "Source File", "Matched Term", "Record ID", "Original DB Row", "NIC", "Passport", "Address", "Phone", 
                  "Birth Day", "First Name", "Last Name", "Type", "Data ID", "Version Number", "Second Name", 
                  "Third Name", "UN List Type", "Reference Number", "Listed On", "Original Script Name", 
                  "Comments", "Title", "Designation", "Nationality", "Last Updated", "Aliases", 
                  "Place of Birth", "Document Number", "NIC No.", "NIC No. 1", "NIC No. 2", "Passport No.", 
                  "Passport No. 1", "Passport No. 2", "Passport No. 3", "Passport No. 4", "DL/Passport No.", 
                  "Citizenship", "Address (Sri Lanka)", "Address (Foreign)", "Listed on", "Other Information", 
                  "Listed on.1"]
        
        for col_num, header in enumerate(headers, 1):
            cell = manual_sheet.cell(row=1, column=col_num, value=header)
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color="00C3FF", end_color="00C3FF", fill_type="solid")
        
        # Add data rows
        for row_num, record in enumerate(result_data['manual'], 2):
            filename = record.get('Source File', '')
            if filename in all_excel_data:
                record_id = record.get('record_id', '')
                if record_id in record_id_map:
                    file_idx = record_id_map[record_id][1]
                    full_record = all_excel_data[filename].iloc[file_idx].to_dict()
                    for key, value in full_record.items():
                        if key in record:
                            record[key] = value
            
            # Write data to excel
            manual_sheet.cell(row=row_num, column=1, value=record.get('Name', 'Unknown').capitalize())
            manual_sheet.cell(row=row_num, column=2, value=record.get('Source File', 'N/A'))
            manual_sheet.cell(row=row_num, column=3, value=record.get('Matched Term', 'N/A'))
            manual_sheet.cell(row=row_num, column=4, value=record.get('record_id', 'N/A'))
            
            # Add original database row number if available
            original_db_row = "N/A"
            if record_id in record_id_map:
                original_db_row = str(record_id_map[record_id][1] + 1)  # +1 because Excel is 1-indexed
            manual_sheet.cell(row=row_num, column=5, value=original_db_row)
            
            # Map additional field values
            field_mapping = {
                "NIC": 6, "Passport": 7, "Address1": 8, "Phone11": 9, "Birth Day": 10,
                "First Name": 11, "Last Name": 12, "Type": 13, "Data ID": 14, "Version Number": 15,
                "Second Name": 16, "Third Name": 17, "UN List Type": 18, "Reference Number": 19,
                "Listed On": 20, "Original Script Name": 21, "Comments": 22, "Title": 23,
                "Designation": 24, "Nationality": 25, "Last Updated": 26, "Aliases": 27,
                "Place of Birth": 28, "Document Number": 29, "NIC No.": 30, "NIC No. 1": 31,
                "NIC No. 2": 32, "Passport No.": 33, "Passport No. 1": 34, "Passport No. 2": 35,
                "Passport No. 3": 36, "Passport No. 4": 37, "DL/Passport No.": 38, "Citizenship": 39,
                "Address (Sri Lanka)": 40, "Address (Foreign)": 41, "Listed on": 42, 
                "Other Information": 43, "Listed on.1": 44
            }
            
            for field, col_idx in field_mapping.items():
                value = record.get(field, 'N/A')
                if value in ('nan', 'None', ''):
                    value = 'N/A'
                manual_sheet.cell(row=row_num, column=col_idx, value=value)
                
                # Highlight matched term
                matched_term = record.get('Matched Term', '').lower()
                if matched_term and matched_term in str(value).lower():
                    manual_sheet.cell(row=row_num, column=col_idx).font = openpyxl.styles.Font(color="FF0000")
                    
                # Highlight fields from the first database file
                if field in ["Type", "Data ID", "Version Number", "Second Name", "Third Name", 
                             "UN List Type", "Reference Number", "Listed On", "Original Script Name", 
                             "Comments", "Title", "Designation", "Nationality", "Last Updated", 
                             "Aliases", "Place of Birth", "Document Number"]:
                    if value != 'N/A':
                        manual_sheet.cell(row=row_num, column=col_idx).fill = openpyxl.styles.PatternFill(
                            start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                
                # Highlight fields from the second database file
                if field in ["NIC No.", "NIC No. 1", "NIC No. 2", "Passport No.", 
                             "Passport No. 1", "Passport No. 2", "Passport No. 3", "Passport No. 4", 
                             "DL/Passport No.", "Citizenship", "Address (Sri Lanka)", 
                             "Address (Foreign)", "Listed on", "Other Information", "Listed on.1"]:
                    if value != 'N/A':
                        manual_sheet.cell(row=row_num, column=col_idx).fill = openpyxl.styles.PatternFill(
                            start_color="E6F2FF", end_color="E6F2FF", fill_type="solid")
            
    else:
        # Create sheet for CSV entries
        entries_sheet = wb.create_sheet(title="Uploaded Entries")
        
        # Add headers
        entry_headers = ["Entry #", "MID", "NIC", "Passport", "Address1", "Phone11", "Birth Day", "First Name", "Last Name"]
        for col_num, header in enumerate(entry_headers, 1):
            cell = entries_sheet.cell(row=1, column=col_num, value=header)
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color="00C3FF", end_color="00C3FF", fill_type="solid")
        
        # Add data rows
        for row_num, (idx, entry) in enumerate(csv_entries.items(), 2):
            entries_sheet.cell(row=row_num, column=1, value=f"Entry #{int(idx) + 1}")
            entries_sheet.cell(row=row_num, column=2, value=entry.get('MID', 'N/A') if entry.get('MID') not in ['None', 'nan'] else 'N/A')
            entries_sheet.cell(row=row_num, column=3, value=entry.get('NIC', 'N/A') if entry.get('NIC') not in ['None', 'nan'] else 'N/A')
            entries_sheet.cell(row=row_num, column=4, value=entry.get('Passport', 'N/A') if entry.get('Passport') not in ['None', 'nan'] else 'N/A')
            entries_sheet.cell(row=row_num, column=5, value=entry.get('Address1', 'N/A') if entry.get('Address1') not in ['None', 'nan'] else 'N/A')
            entries_sheet.cell(row=row_num, column=6, value=entry.get('Phone11', 'N/A') if entry.get('Phone11') not in ['None', 'nan'] else 'N/A')
            entries_sheet.cell(row=row_num, column=7, value=entry.get('Birth Day', 'N/A') if entry.get('Birth Day') not in ['None', 'nan'] else 'N/A')
            entries_sheet.cell(row=row_num, column=8, value=entry.get('First Name', 'N/A') if entry.get('First Name') not in ['None', 'nan'] else 'N/A')
            entries_sheet.cell(row=row_num, column=9, value=entry.get('Last Name', 'N/A') if entry.get('Last Name') not in ['None', 'nan'] else 'N/A')
        
        # Create sheet for matches
        matches_sheet = wb.create_sheet(title="Matches")
        
        # Add headers
        match_headers = ["Entry #", "Match #", "Name", "Source File", "Matched Term", "Record ID", "Original DB Row", 
                        "NIC", "Passport", "Address", "Phone", "Birth Day", "First Name", "Last Name", 
                        "Type", "Data ID", "Version Number", "Second Name", "Third Name", "UN List Type", 
                        "Reference Number", "Listed On", "Original Script Name", "Comments", "Title", "Designation", 
                        "Nationality", "Last Updated", "Aliases", "Place of Birth", "Document Number", 
                        "NIC No.", "NIC No. 1", "NIC No. 2", "Passport No.", "Passport No. 1", "Passport No. 2", 
                        "Passport No. 3", "Passport No. 4", "DL/Passport No.", "Citizenship", "Address (Sri Lanka)", 
                        "Address (Foreign)", "Listed on", "Other Information", "Listed on.1"]
        
        for col_num, header in enumerate(match_headers, 1):
            cell = matches_sheet.cell(row=1, column=col_num, value=header)
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color="00C3FF", end_color="00C3FF", fill_type="solid")
        
        # Add match data
        row_num = 2
        for idx, matches in result_data.items():
            for match_idx, record in enumerate(matches, 1):
                filename = record.get('Source File', '')
                record_id = record.get('record_id', '')
                
                full_record = {}
                original_db_row = "N/A"
                if record_id in record_id_map:
                    file_idx = record_id_map[record_id][1]
                    original_db_row = str(file_idx + 1)  # +1 because Excel is 1-indexed
                    if filename in all_excel_data:
                        full_record = all_excel_data[filename].iloc[file_idx].to_dict()
                
                # Basic info
                matches_sheet.cell(row=row_num, column=1, value=f"Entry #{int(idx) + 1}")
                matches_sheet.cell(row=row_num, column=2, value=f"Match #{match_idx}")
                matches_sheet.cell(row=row_num, column=3, value=record.get('Name', 'Unknown').capitalize())
                matches_sheet.cell(row=row_num, column=4, value=record.get('Source File', 'N/A'))
                matches_sheet.cell(row=row_num, column=5, value=record.get('Matched Term', 'N/A'))
                matches_sheet.cell(row=row_num, column=6, value=record.get('record_id', 'N/A'))
                matches_sheet.cell(row=row_num, column=7, value=original_db_row)
                
                # Map additional field values
                field_mapping = {
                    "NIC": 8, "Passport": 9, "Address1": 10, "Phone11": 11, "Birth Day": 12,
                    "First Name": 13, "Last Name": 14, "Type": 15, "Data ID": 16, "Version Number": 17,
                    "Second Name": 18, "Third Name": 19, "UN List Type": 20, "Reference Number": 21,
                    "Listed On": 22, "Original Script Name": 23, "Comments": 24, "Title": 25,
                    "Designation": 26, "Nationality": 27, "Last Updated": 28, "Aliases": 29,
                    "Place of Birth": 30, "Document Number": 31, "NIC No.": 32, "NIC No. 1": 33,
                    "NIC No. 2": 34, "Passport No.": 35, "Passport No. 1": 36, "Passport No. 2": 37,
                    "Passport No. 3": 38, "Passport No. 4": 39, "DL/Passport No.": 40, "Citizenship": 41,
                    "Address (Sri Lanka)": 42, "Address (Foreign)": 43, "Listed on": 44, 
                    "Other Information": 45, "Listed on.1": 46
                }
                
                for field, col_idx in field_mapping.items():
                    value = full_record.get(field, record.get(field, 'N/A'))
                    if value in ('nan', 'None', ''):
                        value = 'N/A'
                    matches_sheet.cell(row=row_num, column=col_idx, value=str(value))
                    
                    # Highlight matched term
                    matched_term = record.get('Matched Term', '').lower()
                    if matched_term and matched_term in str(value).lower():
                        matches_sheet.cell(row=row_num, column=col_idx).font = openpyxl.styles.Font(color="FF0000")
                    
                    # Highlight fields from the first database file
                    if field in ["Type", "Data ID", "Version Number", "Second Name", "Third Name", 
                                "UN List Type", "Reference Number", "Listed On", "Original Script Name", 
                                "Comments", "Title", "Designation", "Nationality", "Last Updated", 
                                "Aliases", "Place of Birth", "Document Number"]:
                        if value != 'N/A':
                            matches_sheet.cell(row=row_num, column=col_idx).fill = openpyxl.styles.PatternFill(
                                start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                    
                    # Highlight fields from the second database file
                    if field in ["NIC No.", "NIC No. 1", "NIC No. 2", "Passport No.", 
                                "Passport No. 1", "Passport No. 2", "Passport No. 3", "Passport No. 4", 
                                "DL/Passport No.", "Citizenship", "Address (Sri Lanka)", 
                                "Address (Foreign)", "Listed on", "Other Information", "Listed on.1"]:
                        if value != 'N/A':
                            matches_sheet.cell(row=row_num, column=col_idx).fill = openpyxl.styles.PatternFill(
                                start_color="E6F2FF", end_color="E6F2FF", fill_type="solid")
                
                row_num += 1

    if not result_data or (isinstance(result_data, dict) and not result_data):
        # No matches found
        no_match_sheet = wb.create_sheet(title="No Threats Found")
        no_match_sheet['A1'] = "No threats found in List."
        no_match_sheet['A1'].font = openpyxl.styles.Font(size=14, bold=True, color="00FF00")

    # Add visualization through charts
    chart_sheet = wb.create_sheet(title="Visualizations")
    
    # Chart title
    chart_sheet['A1'] = "ThreatCipher Visualization"
    chart_sheet['A1'].font = openpyxl.styles.Font(size=16, bold=True)
    
    # Prepare term frequency data for visualization
    chart_sheet['A3'] = "Term Frequency Data"
    chart_sheet['A3'].font = openpyxl.styles.Font(size=14, bold=True)
    
    chart_sheet['A4'] = "Term"
    chart_sheet['B4'] = "Count"
    chart_sheet['A4'].font = openpyxl.styles.Font(bold=True)
    chart_sheet['B4'].font = openpyxl.styles.Font(bold=True)
    
    # Insert data for top 10 terms
    row = 5
    top_terms = term_counts.most_common(10)
    for term, count in top_terms:
        chart_sheet[f'A{row}'] = term
        chart_sheet[f'B{row}'] = count
        row += 1
    
    # Create a bar chart for term frequency
    chart = openpyxl.chart.BarChart()
    chart.title = "Top Matched Terms"
    chart.style = 10
    chart.x_axis.title = "Count"
    chart.y_axis.title = "Term"
    
    # Define data and categories
    data = openpyxl.chart.Reference(chart_sheet, min_col=2, min_row=4, max_row=row-1, max_col=2)
    categories = openpyxl.chart.Reference(chart_sheet, min_col=1, min_row=5, max_row=row-1)
    
    # Add data to chart
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    
    # Add the chart to the sheet
    chart_sheet.add_chart(chart, "D3")
    
    # Prepare entry-wise match count data for visualization
    chart_sheet['A20'] = "Entry-wise Match Count"
    chart_sheet['A20'].font = openpyxl.styles.Font(size=14, bold=True)
    
    chart_sheet['A21'] = "Entry"
    chart_sheet['B21'] = "Match Count"
    chart_sheet['A21'].font = openpyxl.styles.Font(bold=True)
    chart_sheet['B21'].font = openpyxl.styles.Font(bold=True)
    
    # Insert data for entries
    row = 22
    if 'manual' in result_data:
        chart_sheet[f'A{row}'] = "Manual Search"
        chart_sheet[f'B{row}'] = len(result_data['manual'])
        row += 1
    else:
        for idx, matches in result_data.items():
            chart_sheet[f'A{row}'] = f"Entry #{int(idx) + 1}"
            chart_sheet[f'B{row}'] = len(matches)
            row += 1
    
    # Create a bar chart for entry-wise match count
    entry_chart = openpyxl.chart.BarChart()
    entry_chart.title = "Match Count by Entry"
    entry_chart.style = 10
    entry_chart.x_axis.title = "Entry"
    entry_chart.y_axis.title = "Match Count"
    
    # Define data and categories
    entry_data = openpyxl.chart.Reference(chart_sheet, min_col=2, min_row=21, max_row=row-1, max_col=2)
    entry_categories = openpyxl.chart.Reference(chart_sheet, min_col=1, min_row=22, max_row=row-1)
    
    # Add data to chart
    entry_chart.add_data(entry_data, titles_from_data=True)
    entry_chart.set_categories(entry_categories)
    
    # Add the chart to the sheet
    chart_sheet.add_chart(entry_chart, "D20")

    # Set column widths
    for sheet in wb.worksheets:
        for column_cells in sheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 30)

    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(
        buffer,
        download_name=f"ThreatCipher_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/analyze_data', methods=['POST'])
def analyze_data():
    data = request.get_json()
    result_data = data.get('result_data', {})
    matched_terms = data.get('matched_terms', [])
    csv_entries = data.get('csv_entries', {})

    # Calculate total matched entries
    total_matches = 0
    if 'manual' in result_data:
        total_matches = len(result_data['manual'])
    else:
        total_matches = sum(len(matches) for matches in result_data.values())

    # Calculate term counts and percentage
    term_counts = Counter(matched_terms)
    
    # Entry-wise analytics
    entry_analytics = []
    total_entries = 0
    
    if 'manual' in result_data:
        # For manual search
        total_entries = 1
        entry_terms = set()
        for record in result_data['manual']:
            term = record.get('Matched Term', '')
            if term:
                entry_terms.add(term)
        
        entry_analytics.append({
            'entry_name': 'Manual Search',
            'match_count': len(result_data['manual']),
            'unique_term_count': len(entry_terms),
            'terms': list(entry_terms),
            'percentage': 100.0,  # Manual search always 100%
            'percentage_formatted': '100.0%'
        })
    else:
        # For file uploads
        total_entries = len(csv_entries)
        
        # Calculate average matches
        avg_matches = total_matches / total_entries if total_entries > 0 else 0
        
        # Entry-wise analytics
        for idx, matches in result_data.items():
            entry_idx = int(idx)
            entry_name = f"Entry #{entry_idx + 1}"
            
            # Extract entry details
            entry_details = csv_entries.get(str(entry_idx), {})
            name = f"{entry_details.get('First Name', '')} {entry_details.get('Last Name', '')}".strip()
            if name:
                entry_name += f" ({name})"
            
            # Collect unique terms for this entry
            entry_terms = set()
            for match in matches:
                term = match.get('Matched Term', '')
                if term:
                    entry_terms.add(term)
            
            # Calculate percentage of total matches
            percentage = (len(matches) / total_matches * 100) if total_matches > 0 else 0
            
            entry_analytics.append({
                'entry_idx': entry_idx,
                'entry_name': entry_name,
                'match_count': len(matches),
                'unique_term_count': len(entry_terms),
                'terms': list(entry_terms),
                'percentage': percentage,
                'percentage_formatted': f"{percentage:.1f}%"
            })

    # Sort entries by match count (descending)
    entry_analytics.sort(key=lambda x: x['match_count'], reverse=True)
    
    # Get most frequent terms (top 10 for visualization)
    most_common_terms = term_counts.most_common(10)
    terms = [term for term, count in most_common_terms]
    counts = [count for term, count in most_common_terms]
    
    # Calculate term percentages
    term_percentages = {}
    for term, count in term_counts.items():
        percentage = (count / total_matches * 100) if total_matches > 0 else 0
        term_percentages[term] = percentage
    
    # Get top 10 terms by percentage
    top_terms_percent = sorted(
        [(term, percentage) for term, percentage in term_percentages.items()],
        key=lambda x: x[1],
        reverse=True
    )[:10]
    
    percent_terms = [term for term, pct in top_terms_percent]
    percent_values = [pct for term, pct in top_terms_percent]
    
    # Calculate country-based match analysis (if country data available)
    country_analytics = {}
    country_matches = 0
    
    # Keywords that might indicate country information
    country_indicators = [
        'nationality', 'citizenship', 'address (foreign)', 'country', 'place of birth'
    ]
    
    # Extract countries from matches
    if 'manual' in result_data:
        for record in result_data['manual']:
            record_id = record.get('record_id', '')
            if record_id in record_id_map:
                filename, idx = record_id_map[record_id]
                full_record = all_excel_data[filename].iloc[idx].to_dict()
                
                # Check for country information
                for field in country_indicators:
                    if field in full_record and full_record[field] not in ('', 'nan', 'none', 'n/a'):
                        country = full_record[field].strip().title()
                        if country:
                            country_analytics[country] = country_analytics.get(country, 0) + 1
                            country_matches += 1
    else:
        for idx, matches in result_data.items():
            for match in matches:
                record_id = match.get('record_id', '')
                if record_id in record_id_map:
                    filename, idx = record_id_map[record_id]
                    full_record = all_excel_data[filename].iloc[idx].to_dict()
                    
                    # Check for country information
                    for field in country_indicators:
                        if field in full_record and full_record[field] not in ('', 'nan', 'none', 'n/a'):
                            country = full_record[field].strip().title()
                            if country:
                                country_analytics[country] = country_analytics.get(country, 0) + 1
                                country_matches += 1
    
    # Sort countries by count
    sorted_countries = sorted(
        [(country, count) for country, count in country_analytics.items()],
        key=lambda x: x[1],
        reverse=True
    )[:10]  # Top 10 countries
    
    countries = [country for country, count in sorted_countries]
    country_counts = [count for country, count in sorted_countries]
    
    # Calculate country percentages
    country_percentages = []
    for country, count in sorted_countries:
        percentage = (count / country_matches * 100) if country_matches > 0 else 0
        country_percentages.append(percentage)
    
    # Calculate risk scoring (simple model)
    risk_score = 0
    risk_factors = []
    
    if total_matches > 0:
        # Base risk on total match count
        if total_matches > 10:
            risk_score += 30
            risk_factors.append("High number of matches")
        elif total_matches > 5:
            risk_score += 20
            risk_factors.append("Moderate number of matches")
        else:
            risk_score += 10
            risk_factors.append("Low number of matches")
        
        # Increase risk if specific high-risk countries are found
        high_risk_countries = ["Afghanistan", "Iraq", "Iran", "Syria", "North Korea", "Yemen", "Libya", "Sudan"]
        for country in countries:
            if country in high_risk_countries:
                risk_score += 20
                risk_factors.append(f"Matches from high-risk country: {country}")
                break
        
        # Increase risk if match rate is high for uploaded entries
        if total_entries > 1:  # Only for file uploads
            match_rate = total_matches / total_entries
            if match_rate > 0.5:  # More than 50% of entries have matches
                risk_score += 25
                risk_factors.append("High percentage of entries have matches")
            elif match_rate > 0.2:  # More than 20% of entries have matches
                risk_score += 15
                risk_factors.append("Moderate percentage of entries have matches")
    
    # Cap the risk score at 100
    risk_score = min(risk_score, 100)
    
    # Determine risk level
    risk_level = "Low"
    if risk_score >= 70:
        risk_level = "High"
    elif risk_score >= 40:
        risk_level = "Medium"
    
    response = {
        'total_entries': total_entries,
        'total_matches': total_matches,
        'term_counts': dict(term_counts),
        'entry_analytics': entry_analytics,
        'most_common_terms': terms,
        'most_common_counts': counts,
        'term_percentages': term_percentages,
        'percent_terms': percent_terms,
        'percent_values': percent_values,
        'countries': countries,
        'country_counts': country_counts,
        'country_percentages': country_percentages,
        'risk_score': risk_score,
        'risk_level': risk_level,
        'risk_factors': risk_factors,
        'status': 'success'
    }

    return jsonify(response)

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)